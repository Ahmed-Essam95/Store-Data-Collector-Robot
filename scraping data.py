from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import time
import os

stores_sheet_path = r"File Path Xlsx"
stores_workbook = excel.load_workbook(stores_sheet_path)
stores_sheet = stores_workbook["Sheet1"]

from selenium.webdriver.chrome.options import Options

options_list = Options()
options_list.add_argument("--headless")

stores_robot = webdriver.Chrome(options=options_list)
stores_robot.maximize_window()
stores_robot.get(" Web Link ")
hold = WebDriverWait( stores_robot ,30)


stop_pagination = False
page_counter = 0
store_sq = 0


while stop_pagination == False  :
    page_counter += 1
    # Container of the page stores
    page_stores = []

    # Fetch all store cards
    def re_fetch():
        # return hold.until(EC.visibility_of_all_elements_located((By.XPATH,
        #          "XPATH")))

        return hold.until(EC.visibility_of_all_elements_located((By.XPATH,
                 "XPATH")))


    time.sleep(2)

    # Loop on cards one by one
    for each_card in re_fetch() :

        store_sq += 1

        # one list for one store
        one_card_one_list = []

        # Get and append store area
        area = each_card.find_element(By.CSS_SELECTOR, "div.d-flex > a").text
        one_card_one_list.append(area)

        # Get data of type and code than append
        store_type_code = each_card.find_element(By.CSS_SELECTOR, "div.search-description").text

        shop_type = ""
        for letter in store_type_code :
            if letter != "|" :
                shop_type += letter
            elif letter == "|" :
                break

        one_card_one_list.append(shop_type.strip())

        store_code = store_type_code[len(shop_type)+14:]
        one_card_one_list.append(store_code)

        # Get And Append Location Data
        location = each_card.find_element(By.CSS_SELECTOR, "div.row.mt-2 > div:nth-child(1) > div:nth-child(2)").text
        one_card_one_list.append(location)


        # Store Status Closed Or Else ? ( Element Found Or Not )
        try:
            # Try to catch the Element ( Closed element )
            shop_status = each_card.find_element(By.CSS_SELECTOR, "div > div > div.d-flex > div").text
            one_card_one_list.append("Closed")

        except:
            # If No ( Closed ) Element
            one_card_one_list.append("Open")


        # Opening Time
        date_1 = each_card.find_element(By.CSS_SELECTOR, "div.row.mt-2 > div:nth-child(2) > div:nth-child(2) > div:nth-child(1)").text
        one_card_one_list.append(date_1[23:])

        date_2 = each_card.find_element(By.CSS_SELECTOR, "div.row.mt-2 > div:nth-child(2) > div:nth-child(2) > div:nth-child(2)").text
        one_card_one_list.append(date_2[10:])

        date_3 = each_card.find_element(By.CSS_SELECTOR, "div.row.mt-2 > div:nth-child(2) > div:nth-child(2) > div:nth-child(3)").text
        one_card_one_list.append(date_3[8:])


        # Contact Information
        store_mail = each_card.find_element(By.XPATH, ".//span[text() = 'text: ']//following-sibling::span").text
        one_card_one_list.append(store_mail)

        Mobile_Number = each_card.find_element(By.XPATH, ".//span[text() = 'text ']//following-sibling::span").text
        one_card_one_list.append(Mobile_Number)


        manager_mail = each_card.find_element(By.XPATH, ".//span[text() = 'text ']//following-sibling::span").text
        if manager_mail :
            one_card_one_list.append(manager_mail)

        if not manager_mail :
            one_card_one_list.append("No Data")


        # Manager Phone ? ( Element Found Or Not )
        try:
           manager_phone = each_card.find_element(By.XPATH,".//span[text() = 'text: ']//following-sibling::span").text
           one_card_one_list.append(manager_phone)

        except:
            one_card_one_list.append("no data")


        # Append the page sequence
        one_card_one_list.append(page_counter)


        # Append the Store sequence
        one_card_one_list.append(store_sq)


        # Append full card data to the list of container
        page_stores.append(one_card_one_list)


    # Append the List of lists
    for per_store in page_stores :
        stores_sheet.append(per_store)
    stores_workbook.save(stores_sheet_path)


    print(f"Page Num : {page_counter} || Store Counter : {store_sq}")
    print("-"*40)




    # pagination validation
    next_page_button = stores_robot.find_element(By.XPATH, "//div[@class='class name mt-3']//button[2]")

    if next_page_button.get_attribute("disabled") :
        stop_pagination = True

    if not next_page_button.get_attribute("disabled") :
        stop_pagination = False
        next_page_button.click()





print("Exiting....")
stores_robot.quit()
print("Task has been performed.")

